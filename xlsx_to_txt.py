import csv
from openpyxl import load_workbook

wb = load_workbook("result.xlsx")
ws = wb.active

with open("result.txt", "w", encoding="utf-8", newline='') as f:
    writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)
